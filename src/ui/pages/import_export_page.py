"""导入导出页面 — Excel/CSV 导入物料 + 导出"""

import csv
from pathlib import Path

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QGroupBox, QFileDialog, QMessageBox, QProgressBar, QTextEdit,
)
from PyQt6.QtCore import Qt

from src.core import material_service, category_service
from src.db.database import get_session
from src.db.models import Material
from src.ui.widgets.toast import Toast


class ImportExportPage(QWidget):
    """数据管理 - 导入导出"""

    def __init__(self, config: dict, parent=None):
        super().__init__(parent)
        self.config = config
        self._setup_ui()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 12, 16, 12)

        title = QLabel("数据导入导出")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        root.addWidget(title)

        # ── 导出区域 ──
        export_group = QGroupBox("导出物料数据")
        export_layout = QVBoxLayout()

        export_desc = QLabel("将全部物料数据导出为 Excel (.xlsx) 或 CSV (.csv) 文件。")
        export_desc.setWordWrap(True)
        export_layout.addWidget(export_desc)

        export_btns = QHBoxLayout()
        btn_xlsx = QPushButton("导出 Excel")
        btn_xlsx.setObjectName("primaryButton")
        btn_xlsx.clicked.connect(self._export_xlsx)
        export_btns.addWidget(btn_xlsx)

        btn_csv = QPushButton("导出 CSV")
        btn_csv.clicked.connect(self._export_csv)
        export_btns.addWidget(btn_csv)
        export_btns.addStretch()
        export_layout.addLayout(export_btns)
        export_group.setLayout(export_layout)
        root.addWidget(export_group)

        # ── 导入区域 ──
        import_group = QGroupBox("导入物料数据")
        import_layout = QVBoxLayout()

        import_desc = QLabel(
            "从 Excel (.xlsx) 或 CSV (.csv) 文件导入物料。\n"
            "必需列: material_name, unit\n"
            "可选列: model, package_type, specification, warning_threshold, "
            "storage_location, category_path (如: 电阻/贴片电阻), remarks"
        )
        import_desc.setWordWrap(True)
        import_layout.addWidget(import_desc)

        import_btns = QHBoxLayout()
        btn_import = QPushButton("选择文件导入")
        btn_import.setObjectName("primaryButton")
        btn_import.clicked.connect(self._import_file)
        import_btns.addWidget(btn_import)

        btn_template = QPushButton("下载导入模板")
        btn_template.clicked.connect(self._download_template)
        import_btns.addWidget(btn_template)
        import_btns.addStretch()
        import_layout.addLayout(import_btns)

        self._progress = QProgressBar()
        self._progress.setVisible(False)
        import_layout.addWidget(self._progress)

        import_group.setLayout(import_layout)
        root.addWidget(import_group)

        # 日志区域
        self._log = QTextEdit()
        self._log.setReadOnly(True)
        self._log.setMaximumHeight(200)
        self._log.setPlaceholderText("操作日志…")
        root.addWidget(self._log)

        root.addStretch()

    # ── 导出 ─────────────────────────────────
    _EXPORT_HEADERS = [
        "material_name", "model", "package_type",
        "specification", "unit", "current_stock", "warning_threshold",
        "storage_location", "datasheet_path", "remarks",
    ]

    def _get_export_rows(self) -> list[list]:
        with get_session() as s:
            mats = s.query(Material).order_by(Material.material_name).all()
            rows = []
            for m in mats:
                rows.append([getattr(m, h, "") or "" for h in self._EXPORT_HEADERS])
            return rows

    def _export_xlsx(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", "物料数据.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "物料数据"
            ws.append(self._EXPORT_HEADERS)
            for row in self._get_export_rows():
                ws.append(row)
            wb.save(path)
            self._log.append(f"[OK] 已导出 Excel: {path}")
            Toast.show(self, "Excel 导出成功", "success")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 CSV", "物料数据.csv", "CSV Files (*.csv)"
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self._EXPORT_HEADERS)
                writer.writerows(self._get_export_rows())
            self._log.append(f"[OK] 已导出 CSV: {path}")
            Toast.show(self, "CSV 导出成功", "success")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    # ── 导入 ─────────────────────────────────
    def _import_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择导入文件", "",
            "Excel/CSV Files (*.xlsx *.csv);;All Files (*)"
        )
        if not path:
            return

        try:
            if path.endswith(".xlsx"):
                rows = self._read_xlsx(path)
            elif path.endswith(".csv"):
                rows = self._read_csv(path)
            else:
                QMessageBox.warning(self, "格式错误", "仅支持 .xlsx 和 .csv 文件")
                return

            if not rows:
                QMessageBox.warning(self, "导入失败", "文件内容为空")
                return

            self._do_import(rows, path)
        except Exception as e:
            QMessageBox.warning(self, "导入失败", str(e))

    def _read_xlsx(self, path: str) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        result = []
        for row in rows_iter:
            d = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
            result.append(d)
        wb.close()
        return result

    def _read_csv(self, path: str) -> list[dict]:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            return [row for row in reader]

    def _do_import(self, rows: list[dict], source: str):
        self._progress.setVisible(True)
        self._progress.setRange(0, len(rows))
        success = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows):
            self._progress.setValue(i + 1)
            name = row.get("material_name", "").strip()
            unit = row.get("unit", "个").strip()

            if not name:
                skipped += 1
                continue

            # 处理分类路径
            cat_path = row.get("category_path", "").strip()
            category_id = None
            if cat_path:
                try:
                    cat = category_service.ensure_category_path(cat_path)
                    category_id = cat.id
                except Exception:
                    pass

            data = {
                "material_name": name,
                "model": row.get("model", ""),
                "package_type": row.get("package_type", ""),
                "specification": row.get("specification", ""),
                "unit": unit,
                "warning_threshold": int(row.get("warning_threshold", 10) or 10),
                "storage_location": row.get("storage_location", ""),
                "datasheet_path": row.get("datasheet_path", ""),
                "remarks": row.get("remarks", ""),
                "category_id": category_id,
            }

            try:
                material_service.create_material(data)
                success += 1
            except Exception as e:
                errors.append(f"行{i+2} [{name}]: {e}")

        self._progress.setVisible(False)
        msg = f"导入完成: 成功 {success}, 跳过 {skipped}, 失败 {len(errors)}"
        self._log.append(f"[导入] {msg} (来源: {source})")
        if errors:
            self._log.append("  失败详情:\n  " + "\n  ".join(errors[:20]))
        Toast.show(self, msg, "success" if not errors else "warning")

    # ── 导入模板 ─────────────────────────────
    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存导入模板", "物料导入模板.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "物料导入模板"
            headers = [
                "material_name", "model", "package_type",
                "specification", "unit", "warning_threshold",
                "storage_location", "category_path", "remarks",
            ]
            ws.append(headers)
            # 示例行
            ws.append([
                "贴片电阻 100Ω", "RC0805FR-07100RL",
                "0805", "100Ω ±1% 1/8W", "个", 10,
                "抽屉A-3", "电阻/贴片电阻", "常用阻值",
            ])
            wb.save(path)
            self._log.append(f"[模板] 已保存: {path}")
            Toast.show(self, "模板已保存", "success")
        except Exception as e:
            QMessageBox.warning(self, "保存失败", str(e))
